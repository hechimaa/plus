from tkinter import filedialog
import os
import openpyxl
import datetime
import tkinter as tk


def hojokin(wid, hig):
    hojod = []
    for i in range(len(wid)):
        if wid[i] * hig[i] / 1000000 < 1.6:
            hojod.append(36000)
        elif 1.6 <= wid[i] * hig[i] / 1000000 < 2.8:
            hojod.append(57000)
        elif 2.8 <= wid[i] * hig[i] / 1000000:
            hojod.append(84000)
    return hojod

def kataban(wid, hig, name):
    hojod = []
    if name == 'uti':
        for i in range(len(wid)):
            if 0.2 <= wid[i] * hig[i] / 1000000 < 1.6:
                hojod.append('005PUHT130JSS')
            elif 1.6 <= wid[i] * hig[i] / 1000000 < 2.8:
                hojod.append('005PUHT130JSM')
            elif 2.8 <= wid[i] * hig[i] / 1000000:
                hojod.append('005PUHT130JSL')
            elif wid[i] * hig[i] / 1000000 < 0.2:
                hojod.append('005PUHT130JSX')

    elif name == 'fix':
        for i in range(len(wid)):
            if 0.2 <= wid[i] * hig[i] / 1000000 < 1.6:
                hojod.append('005PUHH160NSS')
            elif 1.6 <= wid[i] * hig[i] / 1000000 < 2.8:
                hojod.append('005PUHH160NSM')
            elif 2.8 <= wid[i] * hig[i] / 1000000:
                hojod.append('005PUHH160NSL')
            elif wid[i] * hig[i] / 1000000 < 0.2:
                hojod.append('005PUHH160NSX')

    else:
        for i in range(len(wid)):
            if 0.2 <= wid[i] * hig[i] / 1000000 < 1.6:
                hojod.append('005PUHF160NSS')
            elif 1.6 <= wid[i] * hig[i] / 1000000 < 2.8:
                hojod.append('005PUHF160NSM')
            elif 2.8 <= wid[i] * hig[i] / 1000000:
                hojod.append('005PUHF160NSL')
            elif wid[i] * hig[i] / 1000000 < 0.2:
                hojod.append('005PUHF160NSX')

    return hojod

def dateshutoku(co):
    global ws
    exdate = []
    for column in ws.iter_cols(min_col=co, max_col=co, min_row=2):
        for cell in column:
            if cell.value == None:
                pass
            else:
                exdate.append(cell.value)
    return exdate

def fukasiteika(wid, hig, sheet2, houkou, mm):
    teika = []
    for j in range(len(wid)):
        sheet = sheet2 + houkou[j] + '-' + mm[j]
        if wid[j] <= 500:
            col = 1
            teika.append(teikashutokurow(col, hig[j], sheet))
        elif 501 <= wid[j] <= 1000:
            col = 2
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 1001 <= wid[j] <= 1500:
            col = 3
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 1501 <= wid[j] <= 2000:
            col = 4
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 2001 <= wid[j] <= 3000:
            col = 5
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 3001 <= wid[j] <= 4000:
            col = 6
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 4001 <= wid[j] <= 5000:
            col = 7
            teika.append(teikashutokurow(col, hig[j], sheet))

        else:
            teika.append(0)
    return teika

def fukasiteikarow(col, row, sheet):
    global wb3
    ws3 = wb3[sheet]

    if 250 <= row <= 800:
        teika = ws3.cell(row=3, column=col).value
    elif 801 <= row <= 1200:
        teika = ws3.cell(row=4, column=col).value
    elif 1201 <= row <= 1400:
        teika = ws3.cell(row=5, column=col).value
    elif 1401 <= row <= 1800:
        teika = ws3.cell(row=6, column=col).value
    elif 1801 <= row <= 2200:
        teika = ws3.cell(row=7, column=col).value
    elif 2201 <= row <= 2450:
        teika = ws3.cell(row=8, column=col).value

    return teika


def utimadoteika(wid, hig, sheet):
    teika = []
    sheet = sheet
    for j in range(len(wid)):
        if 270 <= wid[j] <= 500:
            col = 1
            teika.append(utimadorow(col, hig[j], sheet))
        elif 501 <= wid[j] <= 800:
            col = 2
            teika.append(utimadorow(col, hig[j], sheet))

        else:
            teika.append(0)

    return teika

def utimadorow(col, row, sheet):
    global wb3
    ws3 = wb3[sheet]

    if 434 <= row <= 800:
        teika = ws3.cell(row=3, column=col).value
    elif 801 <= row <= 1200:
        teika = ws3.cell(row=4, column=col).value
    elif 1201 <= row <= 1400:
        teika = ws3.cell(row=5, column=col).value
    elif 1401 <= row <= 1560:
        teika = ws3.cell(row=6, column=col).value
    else:
        teika = 0

    return teika


def teikashutoku(wid, hig, sheet):
    teika = []
    sheet = sheet
    for j in range(len(wid)):
        if wid[j] <= 500:
            col = 1
            teika.append(teikashutokurow(col, hig[j], sheet))
        elif 501 <= wid[j] <= 1000:
            col = 2
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 1001 <= wid[j] <= 1500:
            col = 3
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 1501 <= wid[j] <= 2000:
            col = 4
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 2001 <= wid[j] <= 3000:
            col = 5
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 3001 <= wid[j] <= 4000:
            col = 6
            teika.append(teikashutokurow(col, hig[j], sheet))

        elif 4001 <= wid[j] <= 5000:
            col = 7
            teika.append(teikashutokurow(col, hig[j], sheet))

        else:
            teika.append(0)
    return teika

def teikashutokurow(col, row, sheet):
    global wb3
    ws3 = wb3[sheet]

    if 250 <= row <= 800:
        teika = ws3.cell(row=3, column=col).value
    elif 801 <= row <= 1200:
        teika = ws3.cell(row=4, column=col).value
    elif 1201 <= row <= 1400:
        teika = ws3.cell(row=5, column=col).value
    elif 1401 <= row <= 1800:
        teika = ws3.cell(row=6, column=col).value
    elif 1801 <= row <= 2200:
        teika = ws3.cell(row=7, column=col).value
    elif 2201 <= row <= 2450:
        teika = ws3.cell(row=8, column=col).value

    return teika

def kakikomi(teika, wid, hig, hojo, kake, shouhin, kataban, houkou=[], mm=[]):
    count = 0
    global count2
    for p in teika:
        # ws.cell(row=count + 2, column=4).value = p
        # ws.cell(row=count + 2, column=5).value = hojo[count]

        ws2.cell(row=count2 + 14, column=5).value = wid[count]
        ws2.cell(row=count2 + 14, column=6).value = hig[count]

        if p * kake <= hojo[count]:
            ws2.cell(row=count2 + 14, column=9).value = hojo[count]
        else:
            ws2.cell(row=count2 + 14, column=9).value = p * kake
        ws2.cell(row=count2 + 14, column=11).value = hojo[count]


        keisan.cell(row=count2 + 2, column=10).value = hojo[count]
        keisan.cell(row=count2 + 2, column=4).value = p

        keisan.cell(row=count2 + 2, column=2).value = wid[count]
        keisan.cell(row=count2 + 2, column=3).value = hig[count]


        ws4.cell(row=count2 + 13, column=5).value = wid[count]
        ws4.cell(row=count2 + 13, column=6).value = hig[count]
        ws4.cell(row=count2 + 13, column=9).value = '断熱'
        ws4.cell(row=count2 + 13, column=10).value = 1

        if shouhin == 'ふかし枠':
            ws2.cell(row=count2 + 14, column=2).value = shouhin + houkou[count] + '方 ' + mm[count] + 'mm'
            keisan.cell(row=count2 + 2, column=1).value = shouhin + houkou[count] + '方 ' + mm[count] + 'mm'
            ws4.cell(row=count2 + 13, column=2).value = shouhin + houkou[count] + '方 ' + mm[count] + 'mm'

        else:
            ws2.cell(row=count2 + 14, column=2).value = shouhin
            keisan.cell(row=count2 + 2, column=1).value = shouhin
            ws4.cell(row=count2 + 13, column=2).value = shouhin + ' アルゴン・LOWE・アルミスペーサー'

        ws5.cell(row=count2 + 2, column=3).value = kataban[count]





        count = count + 1
        count2 = count2 + 1


before = datetime.datetime.now()
now = before.strftime('%Y年%m月%d日 %H時%M分')

calentpath = os.getcwd()
print(calentpath)

# 名前入力
# メインウィンドウを作成
baseGround = tk.Tk()
# ウィンドウのサイズを設定
baseGround.geometry('500x300')
# 画面タイトル
baseGround.title('入力フォーム')

# ラベル
label1 = tk.Label(text='名前')
label1.place(x=30, y=70)

label2 = tk.Label(text='施工費')
label2.place(x=30, y=120)

# テキストボックス
textBox1 = tk.Entry(width=40)
textBox1.place(x=30, y=90)

textBox2 = tk.Entry()
textBox2.place(x=30, y=150)

seshuname = 0
kouji = 0

def val():
    # テキストボックスの値を取得
    global seshuname
    global kouji
    print(textBox1.get())
    print(textBox2.get())
    seshuname = textBox1.get()
    kouji = textBox2.get()
    baseGround.destroy()

# ボタンの作成と配置
button = tk.Button(baseGround,
                text = 'OK',
                # クリック時にval()関数を呼ぶ
                command = val
                ).place(x=30, y=180)

baseGround.mainloop()


# 幅と高さを記入したexcelファイルを選択
typ = [('excelファイル','*.xlsx')]
dir = './'
pdf_file_name = filedialog.askopenfilename(filetypes = typ, initialdir = dir)
# save_file = os.path.splitext(pdf_file_name)[0] + '変更済み' + now + '.xlsx'
# save_file2 = os.path.splitext(pdf_file_name)[0] + now + '.xlsx'
excel_file_name = os.path.splitext(pdf_file_name)[0] + '.xlsx'

os.mkdir(calentpath + "\\見積もり\\" + now + seshuname)
os.mkdir(calentpath + "\\見積もり\\" + now + seshuname + '\\写真')
save_file = calentpath + "\\見積もり\\" + "\\" + now + seshuname + "\\" + seshuname + '定価補助金' + now + '.xlsx'
save_file2 = calentpath + "\\見積もり\\" + "\\" + now + seshuname + "\\" + seshuname + now + '.xlsx'
save_file3 = calentpath + "\\見積もり\\" + "\\" + now + seshuname + "\\発注書" + seshuname + now + '.xlsx'

# 幅と高さ取得用のリスト定義
high2m = []
width2m = []
high4m = []
width4m = []

high2sr = []
width2sr = []
high4sr = []
width4sr = []

highunit2 = []
widthunit2 = []

highfix = []
widthfix = []

highuti = []
widthuti = []

highfukasi = []
widthfukasi = []

# excel読み込み
wb = openpyxl.load_workbook(excel_file_name)
ws = wb.active

wb2 = openpyxl.load_workbook(calentpath + "\\原本\\見積もり原本.xlsx")
ws2 = wb2['見積もり']
keisan = wb2['計算表']

wb3 = openpyxl.load_workbook(calentpath + "\\原本\\プラマード価格表.xlsx")

# wblixil = openpyxl.load_workbook(calentpath + "\\原本\\インプラス価格表.xlsx")

# wb4 = openpyxl.load_workbook(calentpath + "\\原本\\発注書.xlsx")
ws4 = wb2['発注書']

ws5 = wb2['型番申請用']

# 補助金額と定価取得用のリスト定義
hojo = []
hojo4m = []
teika2 = []
teika4 = []

hojosr = []
hojo4sr = []
teika2sr = []
teika4sr = []

hojounit2 = []
teikaunit2 = []

hojofix = []
teikafix = []

hojouti = []
teikauti = []

teikafukasi = []

# ふかし枠の種類取得用リスト
fukasihoukou = []
fukasimm = []

# excelから幅と高さを取得
width2m.extend(dateshutoku(2))
high2m.extend(dateshutoku(3))
width4m.extend(dateshutoku(7))
high4m.extend(dateshutoku(8))
width2sr.extend(dateshutoku(12))
high2sr.extend(dateshutoku(13))
width4sr.extend(dateshutoku(17))
high4sr.extend(dateshutoku(18))

widthunit2.extend(dateshutoku(22))
highunit2.extend(dateshutoku(23))

widthfix.extend(dateshutoku(27))
highfix.extend(dateshutoku(28))

widthuti.extend(dateshutoku(32))
highuti.extend(dateshutoku(33))

widthfukasi.extend(dateshutoku(37))
highfukasi.extend(dateshutoku(38))

fukasihoukou.extend(dateshutoku(39))
fukasimm.extend(dateshutoku(40))

for i in range(len(fukasihoukou)):
    fukasihoukou[i] = str(fukasihoukou[i])
    if fukasimm[i] == 2:
        fukasimm[i] = str(25)
    elif fukasimm[i] == 4:
        fukasimm[i] = str(40)


# 補助金額取得
hojo.extend(hojokin(width2m, high2m))
hojo4m.extend(hojokin(width4m, high4m))

hojosr.extend(hojokin(width2sr, high2sr))
hojo4sr.extend(hojokin(width4sr, high4sr))

hojounit2.extend(hojokin(widthunit2, highunit2))

hojofix.extend(hojokin(widthfix, highfix))

hojouti.extend(hojokin(widthuti, highuti))

# 型番取得

kata2 = []
kata4 = []
katasr2 = []
katasr4 = []
kataunit2 = []
katafix = []
katauti = []

kata2.extend(kataban(width2m, high2m, '2'))
kata4.extend(kataban(width4m, high4m, '4'))
katasr2.extend(kataban(width2sr, high2sr, 'sr2'))
katasr4.extend(kataban(width4sr, high4sr, 'sr4'))
kataunit2.extend(kataban(widthunit2, highunit2, 'unit'))
katafix.extend(kataban(widthfix, highfix, 'fix'))
katauti.extend(kataban(widthuti, highuti, 'uti'))

# excel表から定価取得
teika2.extend(teikashutoku(width2m, high2m, 'アルミLOWE2'))
teika4.extend(teikashutoku(width4m, high4m, 'アルミLOWE4'))

teika2sr.extend(teikashutoku(width2sr, high2sr, 'スリ板アルミLOWE2'))
teika4sr.extend(teikashutoku(width4sr, high4sr, 'スリ板アルミLOWE4'))
teikaunit2.extend(teikashutoku(widthunit2, highunit2, 'ユニット用'))

teikafix.extend(teikashutoku(widthfix, highfix, 'FIX'))

teikauti.extend(utimadoteika(widthuti, highuti, '内開き'))

teikafukasi.extend(fukasiteika(widthfukasi, highfukasi, 'ふかし枠', fukasihoukou, fukasimm))

# 補助金ダミー

hojo0 = []
for dm in range(20):
    hojo0.append(0)


# 書き込み

count = 0
count2 = 0

kakikomi(teika2, width2m, high2m, hojo, 0.55, 'YKK 内窓 プラマードU (2枚建)', kata2)
kakikomi(teika4, width4m, high4m, hojo4m, 0.5, 'YKK 内窓 プラマードU (4枚建)', kata4)
kakikomi(teika2sr, width2sr, high2sr, hojosr, 0.55, 'YKK 内窓 プラマードU (2枚建) スリガラス', katasr2)
kakikomi(teika4sr, width4sr, high4sr, hojo4sr, 0.5, 'YKK 内窓 プラマードU (4枚建) スリガラス', katasr4)
kakikomi(teikaunit2, widthunit2, highunit2, hojounit2, 0.55, 'YKK 内窓 プラマードU (2枚建) ユニット用', kataunit2)
kakikomi(teikafix, widthfix, highfix, hojofix, 0.55, 'YKK 内窓 プラマードU FIX窓', katafix)
kakikomi(teikauti, widthuti, highuti, hojouti, 0.55, 'YKK 内窓 プラマードU 内開き窓', katauti)
kakikomi(teikafukasi, widthfukasi, highfukasi, hojo0, 0.5, 'ふかし枠', hojo0, fukasihoukou, fukasimm)




if kouji == None or kouji == 0:
    pass

else:
    ws2.cell(row=count2 + 15, column=9).value = kouji
    ws2.cell(row=count2 + 15, column=2).value = '施工工事費'

ws2.cell(row=3, column=2).value = seshuname + ' 様'
ws2.cell(row=7, column=3).value = seshuname + ' 邸'
ws4.cell(row=7, column=3).value = seshuname + ' 邸'


print(width4sr)
print(high4sr)
print(teika4sr)

# 保存
# wb.save(save_file)
wb2.save(save_file2)
# wb4.save(save_file3)
